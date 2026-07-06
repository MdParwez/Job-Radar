"""
Builds a formatted .xlsx workbook of matched jobs, entirely in memory (no temp
files), for the "Export to Excel" button in the UI.
"""
from datetime import datetime
from io import BytesIO
from typing import List

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo

from app.models.schemas import JobPosting

HEADER_FILL = PatternFill(start_color="1C2140", end_color="1C2140", fill_type="solid")
HEADER_FONT = Font(name="Calibri", bold=True, color="EEF0FB", size=11)
BODY_FONT = Font(name="Calibri", size=10.5)
LINK_FONT = Font(name="Calibri", size=10.5, color="1155CC", underline="single")
THIN_BORDER = Border(bottom=Side(style="thin", color="D9D9D9"))

COLUMNS = [
    ("Title", 30),
    ("Company", 22),
    ("Location", 18),
    ("Source", 14),
    ("Match %", 10),
    ("Match Reason", 40),
    ("Min Exp (yrs)", 12),
    ("Max Exp (yrs)", 12),
    ("Posted", 16),
    ("Salary", 16),
    ("Apply Link", 45),
]


def _hours_ago_to_label(hours_ago) -> str:
    if hours_ago is None:
        return "Recently"
    if hours_ago < 1:
        return "< 1h ago"
    return f"{round(hours_ago)}h ago"


def build_jobs_workbook(jobs: List[JobPosting]) -> BytesIO:
    wb = Workbook()
    ws = wb.active
    ws.title = "Matched Jobs"

    # Header row
    for col_idx, (name, width) in enumerate(COLUMNS, start=1):
        cell = ws.cell(row=1, column=col_idx, value=name)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(vertical="center")
        ws.column_dimensions[get_column_letter(col_idx)].width = width
    ws.freeze_panes = "A2"
    ws.row_dimensions[1].height = 20

    # Data rows
    for row_idx, job in enumerate(jobs, start=2):
        values = [
            job.title,
            job.company,
            job.location or "Remote",
            job.source,
            round(job.match_score, 1) if job.match_score is not None else None,
            job.match_reason or "",
            job.required_experience_min,
            job.required_experience_max,
            _hours_ago_to_label(job.hours_ago),
            job.salary or "",
        ]
        for col_idx, value in enumerate(values, start=1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.font = BODY_FONT
            cell.border = THIN_BORDER
            cell.alignment = Alignment(vertical="center", wrap_text=(col_idx == 6))

        # Apply Link as a real clickable hyperlink, not just plain text
        link_cell = ws.cell(row=row_idx, column=len(COLUMNS), value="Apply →")
        link_cell.hyperlink = job.url
        link_cell.font = LINK_FONT
        link_cell.border = THIN_BORDER

    # Turn the range into a real Excel Table for built-in sort/filter dropdowns
    if jobs:
        last_row = len(jobs) + 1
        last_col_letter = get_column_letter(len(COLUMNS))
        table_ref = f"A1:{last_col_letter}{last_row}"
        table = Table(displayName="MatchedJobs", ref=table_ref)
        table.tableStyleInfo = TableStyleInfo(
            name="TableStyleMedium2", showRowStripes=True, showFirstColumn=False
        )
        ws.add_table(table)

    ws.sheet_view.showGridLines = False

    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer


def export_filename() -> str:
    return f"job-radar-matches-{datetime.now().strftime('%Y%m%d-%H%M')}.xlsx"
